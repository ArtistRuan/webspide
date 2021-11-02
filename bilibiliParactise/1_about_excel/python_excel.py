

from openpyxl import Workbook
from openpyxl import load_workbook

# from openpyxl import Workbook,load_workbook

def dml_excel_sheet():
    # 创建excel实例
    wb = Workbook()
    sheet = wb.active

    print(sheet.title)
    sheet.title = 'python_sheet'

    # 写数据
    sheet["A1"] = '陈丝影'
    sheet["B8"] = '单个数据单独写'
    sheet.append(['多个数据','分别写','你好啊'])
    #
    '''如何在一个excel里面创建多个sheet页面？？？'''
    wb.create_sheet('sheet2', 1)
    wb.create_sheet('sheet1', 2)
    #
    # 保存excel文件
    wb.save("python_excel_text.xlsx")

def open_excel():
    wb = load_workbook('python_excel_text.xlsx')

    # 打印sheet信息
    print(wb.get_sheet_names())
    print(wb.get_sheet_by_name('sheet1'))
    print(wb.get_sheet_by_name('python_sheet'))

    # 拿到sheet指定单元格信息 -----------------------
    print('-------------------------')
    sheet2 = wb.get_sheet_by_name('sheet2')
    print(sheet2['B5'])
    print(sheet2['B8'])
    print(sheet2['B5'].value)
    print(sheet2['B8'].value)

    # 拿到sheet多个单元格信息
    print('------------1、拿到sheet多个单元格信息B列-------------')
    for cell_B in sheet2["B5:B8"]:
        print(cell_B)
        print(cell_B[0].value)

    print('------------2、拿到sheet多个单元格信息A和B列-------------')
    for cell_AB in wb.get_sheet_by_name('python_sheet'):  #拿到python_sheet页
        # 打印A和B列
        print(cell_AB[0].value,cell_AB[1].value)

    print('------------3、拿到sheet多个单元格信息A和B列-------------')
    for row in wb.get_sheet_by_name('python_sheet').iter_rows(min_row=5,max_row=8,max_col=3):
        for cell_col3 in row:
            print(cell_col3.value,end='\t')
        print()


    # 查看指定sheet所有信息 -------------------------
    print('---------查看指定sheet所有信息----------------')
    python_sheet = wb.get_sheet_by_name('python_sheet')
    for row in python_sheet:  # 拿到整个sheet页的每一行row
        for cell in row:
            # print(cell)
            print(cell.value,end='\t')
        print()

    # 2、按列循环
    print('----------2、按列循环----------')
    for col in python_sheet.columns:
        for cell_cols in col:
            print(cell_cols.value,end='\t')
        print()


if __name__ == '__main__':
    dml_excel_sheet()
    # open_excel()